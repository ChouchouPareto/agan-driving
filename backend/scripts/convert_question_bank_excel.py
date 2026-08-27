import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


REQUIRED = {"external_id", "stem", "question_type", "standard_answer", "explanation", "options_json", "knowledge_points_json"}


def convert(source: Path, target: Path, sheet_name: str = "C1题库数据") -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"未找到工作表：{sheet_name}")
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    missing = REQUIRED - set(headers)
    if missing:
        raise ValueError(f"缺少必填列：{', '.join(sorted(missing))}")
    positions = {name: headers.index(name) for name in headers}
    output = []
    for row_number, row in enumerate(rows, 2):
        external_id = str(row[positions["external_id"]] or "").strip()
        if not external_id:
            continue
        try:
            options = json.loads(str(row[positions["options_json"]] or "[]"))
            knowledge_points = json.loads(str(row[positions["knowledge_points_json"]] or "[]"))
        except json.JSONDecodeError as exc:
            raise ValueError(f"第 {row_number} 行 JSON 字段无效：{exc.msg}") from exc
        output.append({
            "external_id": external_id,
            "stem": str(row[positions["stem"]] or "").strip(),
            "question_type": str(row[positions["question_type"]] or "").strip(),
            "options": options,
            "standard_answer": str(row[positions["standard_answer"]] or "").strip(),
            "explanation": str(row[positions["explanation"]] or "").strip(),
            "knowledge_points": knowledge_points,
            "region": str(row[positions.get("region", -1)] or "全国").strip() if "region" in positions else "全国",
            "license_type": str(row[positions.get("license_type", -1)] or "C1").strip() if "license_type" in positions else "C1",
            "image_url": str(row[positions.get("图片URL", -1)] or "").strip() if "图片URL" in positions else "",
            "legal_basis": str(row[positions.get("法条依据", -1)] or "").strip() if "法条依据" in positions else "",
        })
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="将驾校题库 Excel 转为 RAG 导入 JSON")
    parser.add_argument("source", type=Path)
    parser.add_argument("target", type=Path)
    parser.add_argument("--sheet", default="C1题库数据")
    args = parser.parse_args()
    print(f"已转换 {convert(args.source, args.target, args.sheet)} 道题：{args.target}")
